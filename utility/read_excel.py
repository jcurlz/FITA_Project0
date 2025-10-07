import os.path
import time

from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pytest

from utility.CommonMethods import CommonMethods


class Excel():
    def __init__(self, driver :  WebDriver):
         self.driver = driver
         self.wait = WebDriverWait(self.driver, 20)
         self.comm = CommonMethods(self.driver)

    def read_my_excel(self, worksheet):
        filepath = os.path.join(os.getcwd(), "utility", "TestCases.xlsx")
        print("PATH", filepath)
        workbook = load_workbook(filepath)
        sheet = workbook[worksheet]

        header = {}
        for idx_key, cell_val in enumerate(next(sheet.iter_rows(values_only=True))):
            header[cell_val] = idx_key


        for row_cell in sheet.iter_rows(min_row=2, values_only=True):
            test_case = row_cell[header['TestCase']]
            test_description = row_cell[header['TestDescription']]
            step = row_cell[header['TestSteps']]
            action = row_cell[header['Actions']]
            locator_type = row_cell[header['LocatorType']]
            locator_obj = row_cell[header['LocatorObject']]
            try:
                locator = (getattr(By, locator_type), locator_obj)
            except Exception as e:
                pass
            data = row_cell[header['Data']]

            print(f"[{test_case}] Step {step} : {test_description}")

            if action == "ElementClick":
               self.comm.elementClick(locator)

            elif action == "ClearAndEnter":
                self.comm.clearAndEnter(locator, data)

            elif action == "VerifyText":
                self.comm.verify_text(locator, data)
                print(locator)

            elif action == "PrintText":
                self.comm.print_text(locator)

            elif action == "DropdownSelector":
                self.comm.dropdownSelector(locator, data)

            elif action == "Pause":
                time.sleep(50)

            else:
                pass